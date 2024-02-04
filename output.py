from openpyxl import load_workbook
import pandas as pd


def write_to_xlsx(df):
    # The argument is a list with 12 tuples. Each tuples contains the desired data for each month.
    # It creates a xlsx with the desired format.

    # Load the existing workbook
    wb = load_workbook("SII RECIBIDAS PROVEEDORES.xlsx")

    # Select the first worksheet
    ws = wb.worksheets[1]

    # Get the maximum row containing data
    max_row = ws.max_row

    # Delete all rows in a range
    ws.delete_rows(1, max_row)

    headers = df.columns.tolist()
    rows = [headers] + df.values.tolist()
    # Write new data starting from A1
    for row in rows:
        ws.append(row)

    # Save the workbook
    wb.save("SII RECIBIDAS PROVEEDORES.xlsx")
